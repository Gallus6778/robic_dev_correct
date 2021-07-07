import openpyxl
import xlsxwriter
from complaints_correction.correct_odbgprs import Odbgprs
from complaints_correction.correct_apn import Apn
from complaints_correction.correct_qos import Qos
from complaints_correction.correct_gprs_profile import Gprs_profile

class Check_hlr_info_and_provide_decision:
    def __init__(self):
        self.info_parameter = {}
    def main(self):
        xlsx_file = "complaints_internet/storage_xlsx/dataset_internet.xlsx"
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        workbook1 = xlsxwriter.Workbook(xlsx_file)
        worksheet1 = workbook1.add_worksheet()

        # Obtenir les valeurs de la derniere ligne du dataset
        subscriber_info = {
            "imsi": sheet["A" + str(sheet.max_row)].value,
            "encKey": sheet["B" + str(sheet.max_row)].value,
            "algoId": sheet["C" + str(sheet.max_row)].value,
            "kdbId": sheet["D" + str(sheet.max_row)].value,
            "acsub": sheet["E" + str(sheet.max_row)].value,
            "imsiActive": sheet["F" + str(sheet.max_row)].value,
            "accTypeGSM": sheet["G" + str(sheet.max_row)].value,
            "accTypeGERAN": sheet["H" + str(sheet.max_row)].value,
            "accTypeUTRAN": sheet["I" + str(sheet.max_row)].value,
            "odboc": sheet["J" + str(sheet.max_row)].value,
            "odbic": sheet["K" + str(sheet.max_row)].value,
            "odbr": sheet["L" + str(sheet.max_row)].value,
            "odboprc": sheet["M" + str(sheet.max_row)].value,
            "odbssm": sheet["N" + str(sheet.max_row)].value,
            "odbgprs": sheet["O" + str(sheet.max_row)].value,
            "odbsci": sheet["P" + str(sheet.max_row)].value,
            "isActiveIMSI": sheet["Q" + str(sheet.max_row)].value,
            "msisdn": sheet["R" + str(sheet.max_row)].value,
            "actIMSIGprs": sheet["S" + str(sheet.max_row)].value,
            "obGprs": sheet["T" + str(sheet.max_row)].value,
            "qosProfile": sheet["U" + str(sheet.max_row)].value,
            "refPdpContextName": sheet["V" + str(sheet.max_row)].value,
            "imeisv": sheet["W" + str(sheet.max_row)].value,
            "ldapResponse": sheet["X" + str(sheet.max_row)].value,
            "Targets": sheet["Y" + str(sheet.max_row)].value}

        list_apn = ["n-internet","n-connect","n-est","n-cen","n-ada","n-nor","n-sud","n-out","n-nwt","n-exn","n-lit","n-swt"]

        if subscriber_info["ldapResponse"] == 'None':

            if subscriber_info["odbgprs"] != "0" or subscriber_info["odbgprs"] == "None":
                if subscriber_info["odbgprs"] != "0" and subscriber_info["odbgprs"] != "None":
                    Odbgprs(msisdn=subscriber_info['msisdn']).main()
                    self.info_parameter["odbgprs"] = "barring_gprs solved"
                elif subscriber_info["odbgprs"] == "None":
                    Gprs_profile(msisdn=subscriber_info['msisdn']).main()
                    self.info_parameter["odbgprs"] = "barring_gprs_not_defined -> solved"
            #
            if subscriber_info["refPdpContextName"] == "None" or subscriber_info["refPdpContextName"] not in list_apn:
                if subscriber_info["refPdpContextName"] == "None":
                    Gprs_profile(msisdn=subscriber_info['msisdn']).main()
                    self.info_parameter["refPdpContextName"] = "APN is not defined -> solved"
                elif subscriber_info["refPdpContextName"] not in list_apn:
                    if subscriber_info["refPdpContextName"] == "b-connect":
                        self.info_parameter["refPdpContextName"] = "APN=b-connect, pre-configured sim, unable to access to internet"
                    elif subscriber_info["refPdpContextName"] == "n-wap":
                        Apn(msisdn=subscriber_info['msisdn']).main()
                        self.info_parameter["refPdpContextName"] = "APN=n-wap, change to n-internet -> solved"
                    elif subscriber_info["refPdpContextName"] == "n-mms":
                        Apn(msisdn=subscriber_info['msisdn']).main()
                        self.info_parameter["refPdpContextName"] = "APN=n-mms, change to n-internet -> solved"
                    else :
                        self.info_parameter["refPdpContextName"] = "APN is "+ subscriber_info["refPdpContextName"] + ", contact assistant"
            #
            if subscriber_info["qosProfile"] == "None" or subscriber_info['qosProfile'] not in ['QoS-R7', 'GOLD', 'SILVER', 'COPPER']:
                if subscriber_info["qosProfile"] == "None":
                    Gprs_profile(msisdn=subscriber_info['msisdn']).main()
                    self.info_parameter["qosProfile"] = "QoS is not defined"
                elif subscriber_info['qosProfile'] not in ['QoS-R7', 'GOLD', 'SILVER', 'COPPER']:
                    Qos(msisdn=subscriber_info['msisdn']).main()
                    self.info_parameter["qosProfile"] = "QoS " + subscriber_info['qosProfile'] + " is not good -> solved"
            if subscriber_info['actIMSIGprs'] == "false":
                self.info_parameter["actIMSIGprs"] = "actIMSIGprs is " + subscriber_info['actIMSIGprs'] + ". Restart phone"
            if subscriber_info['isActiveIMSI'] == "false":
                self.info_parameter["isActiveIMSI"] = "isActiveIMSI is " + subscriber_info['isActiveIMSI'] + ". Restart phone"

            #
            if (subscriber_info['qosProfile'] in ['QoS-R7', 'GOLD', 'SILVER', 'COPPER']) and (subscriber_info["refPdpContextName"] in list_apn) and (subscriber_info["odbgprs"] == "0") and (subscriber_info['actIMSIGprs'] == "true"):
                self.info_parameter["result"] = "no problem found"

            # Ecrire l'action a mener dans le fichier dataset_internet.xlsx
            values = ''
            for keys, value in self.info_parameter.items():
                values = values + keys + ":" + value + ";"

            sheet.cell(row=sheet.max_row, column=sheet.max_column, value=values)
            wb_obj.save(filename=xlsx_file)
            return self.info_parameter
        elif subscriber_info["ldapResponse"] != 'None':
            self.info_parameter["ldapResponse"] = "Unknow Subscriber"

            #
            values = ''
            for keys, value in self.info_parameter.items():
                values = values + keys + ":" + value + ";"

            sheet.cell(row=sheet.max_row, column=sheet.max_column, value=values)
            wb_obj.save(filename=xlsx_file)
            return self.info_parameter
        else:
            pass

if __name__ == "__main__":
    subscriber_info = Check_hlr_info_and_provide_decision()
    info_parameter = subscriber_info.main()

    print(info_parameter)
        # ----------------- Consultation parametres HLR ----------------------