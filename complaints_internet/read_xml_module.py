import xml.etree.cElementTree as ET
import openpyxl

class Read_request_by_msisdn:
    def __init__(self, xml_reveive_from_hlr,msisdn_info_results):
        self.xml_reveive_from_hlr = xml_reveive_from_hlr
        self.msisdn_info_results = msisdn_info_results

    def put_data_in_dataset(self):

        """ Return dictionnary of xml value"""

        # Extraction du contenu du fichier xml
        tree = ET.ElementTree(file=self.xml_reveive_from_hlr)
        root = tree.getroot()
        inc = 1
        for books in root.findall('.//'):

            if (books.tag == 'imsi'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'encKey'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'algoId'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'kdbId'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'acsub'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'imsiActive'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'accTypeGSM'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'accTypeGERAN'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'accTypeUTRAN'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odboc'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbic'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbr'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odboprc'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbssm'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbgprs'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbsci'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'msisdn'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'isActiveIMSI'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'actIMSIGprs'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'obGprs'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'qosProfile'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'imeisv'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'ldapResponse'):
                self.msisdn_info_results[books.tag] = books.text

            ident = ''
            if (books.tag == 'pdpContext'):
                for attr in books:
                    if inc == 1:
                        if (attr.tag == "id"):
                            ident = attr.text
                        if (attr.tag == "refPdpContextName"):
                            refPdpContextName = attr.text
                            if ident == "1" or ident == "5":
                                self.msisdn_info_results[attr.tag] = refPdpContextName

                    if inc == 2:
                        if (attr.tag == "id"):
                            ident = attr.text
                        if (attr.tag == "refPdpContextName"):
                            refPdpContextName = attr.text
                            if ident == "1" or ident == "5":
                                self.msisdn_info_results[attr.tag] = refPdpContextName
                    if inc == 3:
                        if (attr.tag == "id"):
                            ident = attr.text
                        if (attr.tag == "refPdpContextName"):
                            refPdpContextName = attr.text
                            if ident == "1" or ident == "5":
                                self.msisdn_info_results[attr.tag] = refPdpContextName
                inc += 1

        # --------------------------------------------------------------------------------------------------------------
        # ----------------- collectes des donnees dans le fichier excel -----------------------------------------------#
        # --------------------------------------------------------------------------------------------------------------

        try:
        # -----------------------------------------------------------------------------
        # --------Generation du fichier Excel (Avec lignes et colonnes)----------------
        # -----------------------------------------------------------------------------

            # workbook = openpyxl.Workbook()
            # sheet = workbook.active
            # column = 1
            # for key,values in self.msisdn_info_results.items():
            #     sheet.cell(row=1, column=column, value=key)
            #     row = 2
            #     sheet.cell(row=row, column=column, value=values)
            #     column += 1
            # workbook.save(filename='complaints_internet/storage_xlsx/dataset_internet.xlsx')

        #
        # -----------------------------------------------------------------------------
        # ----------------- Ajout des donnees a la fin du fichier excel----------------
        # -----------------------------------------------------------------------------

            filename = 'complaints_internet/storage_xlsx/dataset_internet.xlsx'
            wb = openpyxl.load_workbook(filename=filename)

            sheet = wb.active
            new_row = []

            for key, values in self.msisdn_info_results.items():
                new_row.append(values)

            sheet.append(new_row)
            wb.save(filename)

        except :
            messageErreur = 'Error -> file not closed:-) You must first closed the "dataset_internet.xlsx" file !'
            return messageErreur



if __name__ == "__main__":
    file = 'soap.xml'
    msisdn_info_results = {}

    Read_request_by_msisdn = Read_request_by_msisdn(file, msisdn_info_results)
    Read_request_by_msisdn.put_data_in_dataset()