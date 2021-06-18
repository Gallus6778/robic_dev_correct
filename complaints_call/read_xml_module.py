import xml.etree.cElementTree as ET
import openpyxl
# from complaints_call.check_msisdn_in_msc_module import Check_msisdn

class Read_xml_for_complaints_call:
    def __init__(self, xml_reveive_from_hlr,msisdn_info_results):
        self.xml_reveive_from_hlr = xml_reveive_from_hlr
        self.msisdn_info_results = msisdn_info_results

    def put_data_in_dataset(self):

        """ Return dictionnary of xml value"""

        # Extraction du contenu du fichier xml
        tree = ET.ElementTree(file=self.xml_reveive_from_hlr)
        root = tree.getroot()
        cfu = 1
        cfb = 1
        cfnrc = 1
        cfnry = 1

        for books in root.findall('.//'):

            if (books.tag == 'imsi'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'imsiActive'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odboc'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbic'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'odbr'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'msisdn'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'isActiveIMSI'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'imeisv'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'ldapResponse'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'vlrIdValid'):
                self.msisdn_info_results[books.tag] = books.text

            if (books.tag == 'isdnNumberOfVLR'):
                self.msisdn_info_results[books.tag] = books.text
                if books.text == "237660001051":
                    self.msisdn_info_results['mss_ip'] = "10.124.208.1"
                elif books.text == "237660001052":
                    self.msisdn_info_results['mss_ip'] = "10.124.208.81"
                elif books.text == "237660002051":
                    self.msisdn_info_results['mss_ip'] = "10.124.140.1"
                elif books.text == "237660002052":
                    self.msisdn_info_results['mss_ip'] = "10.124.148.4"

            if (books.tag == 'cfu'):
                for attr in books:
                    if cfu == 1:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfu'] = attr.text

                    if cfu == 2:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfu'] = attr.text
                cfu += 1

            if (books.tag == 'cfb'):
                for attr in books:
                    if cfb == 1:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfb'] = attr.text

                    if cfb == 2:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfb'] = attr.text
                cfb += 1

            if (books.tag == 'cfnrc'):
                for attr in books:
                    if cfnrc == 1:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfnrc'] = attr.text

                    if cfnrc == 2:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfnrc'] = attr.text
                cfnrc += 1

            if (books.tag == 'cfnry'):
                for attr in books:
                    if cfnry == 1:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfnry'] = attr.text

                    if cfnry == 2:
                        if (attr.tag == "isdnNumber"):
                            self.msisdn_info_results['cfnry'] = attr.text
                cfnry += 1

        # -----------------------------------------------------------------------------
        # ----------------- collectes des donnees dans le fichier excel ---------------
        # -----------------------------------------------------------------------------

        try:
        # -----------------------------------------------------------------------------
        # --------Generation du fichier Excel (Avec lignes et colonnes)----------------
        # -----------------------------------------------------------------------------

            # workbook = openpyxl.Workbook()
            # sheet = workbook.active
            # column = 1
            # for key,values in self.msisdn_info_results.items():
            #     sheet.cell(row=1, column=column, value=key)
            #     row = 2
            #     sheet.cell(row=row, column=column, value=values)
            #     column += 1
            # workbook.save(filename='complaints_call/storage_xlsx/dataset_call.xlsx')

        #
        # -----------------------------------------------------------------------------
        # ----------------- Ajout des donnees a la fin du fichier excel----------------
        # -----------------------------------------------------------------------------

            filename = 'complaints_call/storage_xlsx/dataset_call.xlsx'
            wb = openpyxl.load_workbook(filename=filename)

            sheet = wb.active
            new_row = []

            for key, values in self.msisdn_info_results.items():
                new_row.append(values)

            sheet.append(new_row)
            wb.save(filename)

        except :
            messageErreur = 'Error -> file not closed:-) You must first closed the "dataset_call.xlsx" file !'
            return messageErreur

if __name__ == "__main__":
    file = 'soap.xml'
    msisdn_info_results = {}
    Read_request_by_msisdn = Read_xml_for_complaints_call(file, msisdn_info_results)
    Read_request_by_msisdn.put_data_in_dataset()