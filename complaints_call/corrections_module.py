import openpyxl
import xlsxwriter
from complaints_correction.correct_obic import Obic
from complaints_correction.correct_oboc import Oboc

class Check_hlr_info_and_provide_decision:
    def __init__(self):
        self.info_parameter = {}
    def main(self):
        xlsx_file = "complaints_call/storage_xlsx/dataset_call.xlsx"
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        workbook1 = xlsxwriter.Workbook(xlsx_file)
        worksheet1 = workbook1.add_worksheet()

        # Obtenir les valeurs de la derniere ligne du dataset
        subscriber_info = {
            "imsi": sheet["A" + str(sheet.max_row)].value,
            "imsiActive": sheet["B" + str(sheet.max_row)].value,
            "odboc": sheet["C" + str(sheet.max_row)].value,
            "odbic": sheet["D" + str(sheet.max_row)].value,
            "odbr": sheet["E" + str(sheet.max_row)].value,
            "msisdn": sheet["F" + str(sheet.max_row)].value,
            "isActiveIMSI": sheet["G" + str(sheet.max_row)].value,
            "imeisv": sheet["H" + str(sheet.max_row)].value,
            "ldapResponse": sheet["I" + str(sheet.max_row)].value,
            "vlrIdValid": sheet["J" + str(sheet.max_row)].value,
            "isdnNumberOfVLR": sheet["K" + str(sheet.max_row)].value,
            "mss_ip": sheet["L" + str(sheet.max_row)].value,
            "cfu": sheet["M" + str(sheet.max_row)].value,
            "cfb": sheet["N" + str(sheet.max_row)].value,
            "cfnrc": sheet["O" + str(sheet.max_row)].value,
            "cfnry": sheet["P" + str(sheet.max_row)].value,
            "Targets": sheet["U" + str(sheet.max_row)].value}
        subscriber_exists_in_msc = {
            "YAMS01": sheet["Q" + str(sheet.max_row)].value,
            "YAMS02": sheet["R" + str(sheet.max_row)].value,
            "DOMS01": sheet["S" + str(sheet.max_row)].value,
            "DOMS02": sheet["T" + str(sheet.max_row)].value}
        number_of_msc_where_subscriber_exists = 0
        if subscriber_info["ldapResponse"] == 'None':
            #
            if subscriber_info['isActiveIMSI'] == "true":
                if subscriber_info["odboc"] != "0" or subscriber_info["odboc"] == "None":
                    if subscriber_info["odboc"] != "0" and subscriber_info["odboc"] != "None":
                        Oboc(msisdn=subscriber_info['msisdn']).main()
                        self.info_parameter["odboc"] = "Barring oc solved"
                #
                for keys, value in subscriber_exists_in_msc.items():
                    if value == 'KNOWN SUBSCRIBER':
                        self.info_parameter[keys] = value
                        number_of_msc_where_subscriber_exists +=1
                #
                if number_of_msc_where_subscriber_exists == 0:
                    self.info_parameter['no_exist'] = 'Subscriber exists in any MSS. Restart phone'
                #
                if subscriber_info["odbic"] != "0" or subscriber_info["odbic"] == "None":
                    if subscriber_info["odbic"] != "0" and subscriber_info["odbic"] != "None":
                        Obic(msisdn=subscriber_info['msisdn']).main()
                        self.info_parameter["odbic"] = "Barring ic solved"
                #
                if subscriber_info["odbr"] != "0" or subscriber_info["odbr"] == "None":
                    if subscriber_info["odbr"] != "0" and subscriber_info["odbr"] != "None":
                        self.info_parameter["odbr"] = "Barring for roaming in HLR"
                    elif subscriber_info["odbr"] == "None":
                        self.info_parameter["odbr"] = "Barring for roaming not defined"
                #
                if subscriber_info['cfu'] != "None":
                    self.info_parameter["cfu"] = "cfu is defined to " + subscriber_info['cfu']
                #
                if subscriber_info['cfb'] != "None":
                    self.info_parameter["cfb"] = "cfb is defined to " + subscriber_info['cfb']
                #
                if subscriber_info['cfnrc'] != "None":
                    self.info_parameter["cfnrc"] = "cfnrc is defined to " + subscriber_info['cfnrc']
                #
                if subscriber_info['cfnry'] != "None":
                    self.info_parameter["cfnry"] = "cfnry is defined to " + subscriber_info['cfnry']
                #
            else:
                self.info_parameter["isActiveIMSI"] = "Subscriber is not attached. Try to restart phone"
            #
            if (subscriber_info["odboc"] == "0") and (subscriber_info["odbic"] == "0") and (
                    subscriber_info["odbr"] == "0") and (subscriber_info['isActiveIMSI'] == "true") and subscriber_info[
                'cfnry'] == "None" and subscriber_info['cfnrc'] == "None" and subscriber_info['cfb'] == "None" and \
                    subscriber_info['cfu'] == "None" and number_of_msc_where_subscriber_exists <= 1:
                self.info_parameter["result"] = "ok"


            # Ecrire l'action a mener dans le fichier dataset_call.xlsx
            values = ''
            for keys, value in self.info_parameter.items():
                values = values + keys + ":" + value + ";"

            sheet.cell(row=sheet.max_row, column=sheet.max_column, value=values)
            wb_obj.save(filename=xlsx_file)
            return self.info_parameter
        else:
            self.info_parameter["ldapResponse"] = "Unknow Subscriber in HLR"

            #
            values = ''
            for keys, value in self.info_parameter.items():
                values = values + keys + ":" + value + ";"

            sheet.cell(row=sheet.max_row, column=sheet.max_column, value=values)
            wb_obj.save(filename=xlsx_file)
            return self.info_parameter

if __name__ == "__main__":
    subscriber_info = Check_hlr_info_and_provide_decision()
    info_parameter = subscriber_info.main()

    print(info_parameter)
        # ----------------- Consultation parametres HLR ----------------------